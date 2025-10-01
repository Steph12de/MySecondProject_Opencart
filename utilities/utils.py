import string
import random

from openpyxl import workbook, load_workbook


class Utils:
    def read_data_from_excel(file, sheet):
        wb = load_workbook(file)
        ws = wb[sheet]
        data_list = []

        T_rows = ws.max_row
        T_columns = ws.max_column

        for r in range(2, T_rows + 1):
            row = []
            for c in range(1, T_columns + 1):
                row.append(ws.cell(r, c).value)
            data_list.append(row)
        return data_list

    @staticmethod
    def random_email():
        domains = ["gmail.com", "yahoo.com", "outlook.com"]
        letters = string.ascii_lowercase
        length_of_mail = 10

        # letter_list = [random.choice(letters) for i in range(length_of_mail)]
        letter_list = []
        for i in range(length_of_mail):
            letter = random.choice(letters)
            letter_list.append(letter)
        return f"{''.join(letter_list)}@{random.choice(domains)}"

        # return f"emailtest{random.randint(1, 1000)}@{random.choice(domains)}"

